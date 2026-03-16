#!/usr/bin/env python3
"""
Inspect the second payroll Excel file and print basic structure + a small sample.

This is intended to be run locally in the project workspace and will print:
- sheet names
- sheet dimensions (row/col count)
- the first few rows (including the header row if present)
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Iterable

import openpyxl  # type: ignore


def _format_row(row: Iterable[Any]) -> str:
    return ", ".join("" if v is None else str(v) for v in row)


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="read_excel_two",
        description="Inspect a payroll Excel file and show basic structure / sample rows.",
    )
    parser.add_argument(
        "path",
        nargs="?",
        default=Path(
            r"C:\Users\Mr. Green\Downloads\Phone Link\PAYROL MOCK DATA FEBRUARY  MAIN SALARY,2026.xlsx"
        ),
        type=Path,
        help="Path to the payroll Excel file",
    )

    args = parser.parse_args()
    path: Path = args.path

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print("Loaded workbook:", path)
    print("Sheets:", wb.sheetnames)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"\n--- Sheet: {sheet_name} (rows={max_row}, cols={max_col})")

        # Print the first few rows for a quick glance
        for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(6, max_row), values_only=True), start=1
        ):
            print(f"Row {i}:", _format_row(row))

        # Identify the first non-empty row and treat it as header row
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=min(12, max_row), values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                header_row = row
                break

        if header_row:
            print("Header (first non-empty) :", _format_row(header_row))

        # Print one example data row (after header)
        data_row = None
        for row in ws.iter_rows(min_row=2, max_row=min(51, max_row), values_only=True):
            if any(cell is not None for cell in row):
                data_row = row
                break

        if data_row:
            print("Sample data row      :", _format_row(data_row))

    print("\nDone.")


if __name__ == "__main__":
    main()
