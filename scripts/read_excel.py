from pathlib import Path
from typing import Any, Iterable

import openpyxl

# Path to the supplied Excel file
EXCEL_PATH = Path(
    r"C:\Users\Mr. Green\Downloads\Phone Link\NO-DATA PAYROL MOCK DATA FEBRUARY  MAIN SALARY, 2026.xlsx"
)


def _format_row(row: Iterable[Any]) -> str:
    return ", ".join("" if v is None else str(v) for v in row)


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print("Loaded workbook:", EXCEL_PATH)
    print("Sheets:", wb.sheetnames)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        print(f"\n--- Sheet: {sheet_name} (rows={max_row}, cols={max_col})")

        # Print the first few rows for a quick glance
        for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(6, max_row), values_only=True), start=1
        ):
            print(f"Row {i}: {_format_row(row)}")

        # Find the first non-empty row as a likely header row
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=min(10, max_row), values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                header_row = row
                break

        if header_row:
            print("Header row (first non-empty):", _format_row(header_row))

        # Print a single sample data row (if available)
        data_rows = list(
            ws.iter_rows(min_row=2, max_row=min(51, max_row), values_only=True)
        )
        if data_rows:
            print("Sample data row:", _format_row(data_rows[0]))

    print("\nDone.")


if __name__ == "__main__":
    main()
