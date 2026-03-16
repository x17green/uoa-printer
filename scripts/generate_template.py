#!/usr/bin/env python3
"""
Generate a cleaned payroll template workbook from the master payroll sheets.

This script reads the two master sheets ("TEACHING STAFF" and "NON TEACHING STAFF")
from the supplied payroll workbook and writes a new workbook that:

  - Adds a required `DOB` column (blank placeholder; required for future ID generation)
  - Adds a generated `STAFF_ID` (8 chars, deterministic-ish)
  - Preserves all original columns
  - Computes and adds validation columns:
      - CALC_NET_FROM_GROSS (gross - deductions)
      - NET_DIFFERENCE (reported net - calculated net)
  - Ensures the resulting workbook is a stable “template” that can be used for
    upload and validation in the system.

Usage:
  python scripts/generate_template.py
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, List, Optional

import openpyxl
from openpyxl import Workbook

INPUT_PATH = Path("docs/data/PAYROL MOCK DATA FEBRUARY  MAIN SALARY,2026.xlsx")
OUTPUT_PATH = Path("docs/data/PAYROLL_TEMPLATE_WITH_DOB_AND_CALCULATIONS.xlsx")

MASTER_SHEETS = ["TEACHING STAFF", "NON TEACHING STAFF"]

# Headers we will look for in each worksheet (case-insensitive).
GROSS_HEADERS = {"GROSS PAY", "GROSS PAYBEFORE", "GROSS PAY BEFORE"}
DEDUCTION_HEADERS = {"TOTAL DEDUCTIONS", "TOT DEDUCT", "TOTAL DEDUCTION"}
NET_HEADERS = {"NET PAY", "NETPAY", "NET_PAY"}


def normalize_header(header: Any) -> str:
    if header is None:
        return ""
    return str(header).strip().upper()


def _find_column_index(header_row: List[Any], candidates: set[str]) -> Optional[int]:
    """
    Return the index (0-based) of the first header cell that matches any candidate.
    """
    for idx, cell in enumerate(header_row):
        if normalize_header(cell) in candidates:
            return idx
    return None


def _as_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return float(value)
    except Exception:
        return 0.0


def _generate_staff_id(name: str, row_index: int) -> str:
    """
    Generate an 8-character staff ID from a name and row index.
    This is deterministic and stable for the same name, but not immutable.
    """
    normalized = "".join([c for c in name.upper() if c.isalpha()])
    normalized = normalized or "STAFF"
    chunk = (normalized + "XXXX")[:4]
    suffix = f"{row_index:04d}"[-4:]
    candidate = f"{chunk}{suffix}"
    return candidate[:8]


def _find_header_row(rows: List[tuple]) -> int:
    """
    Find the first row that contains an obvious header (a row containing "STAFF" and "NAME").
    Fallback to the first row.
    """
    for i, row in enumerate(rows):
        joined = " ".join([str(c or "").upper() for c in row])
        if "STAFF" in joined and "NAME" in joined:
            return i
    return 0


def process_sheet(wb_in: openpyxl.Workbook, sheet_name: str, wb_out: Workbook) -> None:
    ws_in = wb_in[sheet_name]
    ws_out = wb_out.create_sheet(sheet_name)

    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        return

    header_index = _find_header_row(rows)
    header_row = list(rows[header_index])

    # Add our template columns (STAFF_ID, DOB) before the original headers
    output_header = (
        ["STAFF_ID", "DOB"]
        + [str(c or "") for c in header_row]
        + [
            "CALC_NET_FROM_GROSS",
            "NET_DIFFERENCE",
        ]
    )
    ws_out.append(output_header)

    # Find the important columns
    gross_idx = _find_column_index(header_row, GROSS_HEADERS)
    deduction_idx = _find_column_index(header_row, DEDUCTION_HEADERS)
    net_idx = _find_column_index(header_row, NET_HEADERS)

    for row_i, row in enumerate(rows[header_index + 1 :], start=1):
        # Skip fully empty rows
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        # Extract a name candidate for staff ID generation
        name_val = ""
        for col in row:
            if isinstance(col, str) and "STAFF" in col.upper() and len(col.strip()) > 1:
                name_val = col.strip()
                break
        if not name_val:
            for col in row:
                if isinstance(col, str) and col.strip():
                    name_val = col.strip()
                    break

        staff_id = _generate_staff_id(name_val or "UNKNOWN", row_i)

        gross = (
            _as_number(row[gross_idx])
            if gross_idx is not None and gross_idx < len(row)
            else 0.0
        )
        deductions = (
            _as_number(row[deduction_idx])
            if deduction_idx is not None and deduction_idx < len(row)
            else 0.0
        )
        net = (
            _as_number(row[net_idx])
            if net_idx is not None and net_idx < len(row)
            else 0.0
        )

        calc_net = gross - deductions
        net_diff = net - calc_net

        ws_out.append([staff_id, "", *row, calc_net, net_diff])


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_PATH}")

    wb_in = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sheet in MASTER_SHEETS:
        if sheet in wb_in.sheetnames:
            process_sheet(wb_in, sheet, wb_out)

    wb_out.save(OUTPUT_PATH)
    print(f"Generated template workbook: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
